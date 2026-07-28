from pathlib import Path
from datetime import datetime
import csv
import math
import re
import sys
import time
import tkinter as tk
from tkinter import filedialog

from tabulate import tabulate

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


# ==============================================================================
# CONFIGURATION
# ==============================================================================

TARGET_CHANNELS = [                # Every channel listed here is analyzed
    "Att_Channel-1",                # in an independent pass, using the same
    "Att_Channel-3",                # processing pipeline. Add/remove entries
]                                   # here to change which channels are reported.

CHANNEL_UPPER_LIMITS = {            # Highest threshold (dB), per channel.
    "Att_Channel-1": 36.00,         # Used when a detected channel matches
    "Att_Channel-3": 31.00,         # one of these known entries.
}

DEFAULT_UPPER_LIMIT = 60.00         # Fallback ceiling for any detected
                                    # channel with no entry above — keeps
                                    # the engine compatible with datasets
                                    # (e.g. legacy) whose channel set isn't
                                    # pre-listed here.

LOWER_LIMIT = 1.00                 # Lowest threshold (dB) — shared by all channels
STEP_SIZE   = 0.10                 # Threshold step size (dB) — shared by all channels

MONTH_ORDER = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ==============================================================================
# FOLDER SELECTION
# ==============================================================================

def select_year_folder() -> str:
    """
    Opens a Windows folder selection dialog and returns the path chosen by
    the user (expected to be a Processed_Data/<year> folder).  If the user
    cancels or closes the dialog, prints a message and exits the program.
    """
    root = tk.Tk()
    root.withdraw()                         # hide the empty root window

    folder_path = filedialog.askdirectory(
        title="Select Year Folder",
    )

    root.destroy()

    if not folder_path:
        print("No folder selected. Exiting.")
        sys.exit(0)

    return folder_path


# ==============================================================================
# DISCOVERY
# ==============================================================================

def find_months(year_folder: str):
    """
    Returns a sorted list of month folder Paths inside the selected year
    folder (e.g. Processed_Data/2020/January_2020, .../February_2020).
    """
    year_path = Path(year_folder)

    if not year_path.exists():
        print(f"Warning: '{year_path}' does not exist. Nothing to analyze.")
        return []

    return sorted(
        folder
        for folder in year_path.iterdir()
        if folder.is_dir()
    )


# Matches both legacy (Attenuation_NAR_D_M_YYYY.txt) and current
# (Attenuation_NARL_D_M_YYYY.txt) processed filenames — the "L?" makes the
# "L" optional so a single pattern covers both dataset generations.
ATTENUATION_FILENAME_PATTERN = re.compile(
    r"^Attenuation_NARL?_\d{1,2}_\d{1,2}_\d{4}\.txt$"
)


def find_attenuation_files(month_folder):
    """
    Recursively finds every processed attenuation file inside a month
    folder (one level down, inside each rainy-day folder), matching both
    the legacy naming convention (Attenuation_NAR_D_M_YYYY.txt) and the
    current naming convention (Attenuation_NARL_D_M_YYYY.txt).  PNG files
    are ignored entirely.

    Returns a sorted list of file Paths.
    """
    month_path = Path(month_folder)

    return sorted(
        f for f in month_path.rglob("*.txt")
        if ATTENUATION_FILENAME_PATTERN.match(f.name)
    )


def detect_available_channels(year_folder: str) -> list:
    """
    Scans every attenuation file found anywhere within the selected year
    folder and returns the sorted list of Att_Channel-N columns that
    actually exist in the processed data (union of headers across every
    file found).

    This lets the engine automatically adapt to whichever dataset
    produced the processed files — Current Dataset (multiple channels)
    or Legacy Dataset (a single channel) — without asking the user or
    hardcoding a dataset type anywhere.
    """
    channels_found = set()

    for month_folder in find_months(year_folder):
        for file_path in find_attenuation_files(month_folder):
            try:
                with open(file_path, "r", newline="") as f:
                    reader = csv.DictReader(f, delimiter="\t")
                    fieldnames = reader.fieldnames or []
            except (OSError, csv.Error):
                continue

            for name in fieldnames:
                if re.match(r"^Att_Channel-\d+$", name):
                    channels_found.add(name)

    def _channel_sort_key(name: str) -> int:
        match = re.search(r"(\d+)$", name)
        return int(match.group(1)) if match else 0

    return sorted(channels_found, key=_channel_sort_key)


def _month_label(month_folder_name: str) -> str:
    """
    Maps a month folder name (e.g. "January_2020") to its short calendar
    name (e.g. "January") used for table grouping and ordering.

    Falls back to the raw folder name if no known month name is found.
    """
    for month in MONTH_ORDER:
        if month_folder_name.lower().startswith(month.lower()):
            return month

    return month_folder_name


# ==============================================================================
# ANALYSIS
# ==============================================================================

def _build_thresholds(upper_limit: float):
    """
    Builds the list of thresholds from LOWER_LIMIT to upper_limit
    (inclusive) in STEP_SIZE increments.  Values are rounded to 2 decimal
    places to avoid floating-point drift.
    """
    thresholds = []
    n_steps = round((upper_limit - LOWER_LIMIT) / STEP_SIZE)

    for i in range(n_steps + 1):
        value = LOWER_LIMIT + (i * STEP_SIZE)
        thresholds.append(round(value, 2))

    return thresholds


def _read_channel_values(file_path, channel_name):
    """
    Reads a single attenuation file and returns a list of valid
    float values for the given channel_name (NaN and +/- infinity excluded).

    Skips and warns on unreadable files or missing columns.  Skips
    individual malformed rows without aborting the rest of the file.
    """
    values = []

    try:
        with open(file_path, "r", newline="") as f:
            reader = csv.DictReader(f, delimiter="\t")

            if channel_name not in (reader.fieldnames or []):
                print(
                    f"Warning: '{channel_name}' column not found in "
                    f"{file_path} — skipping file."
                )
                return values

            for row in reader:
                try:
                    value = float(row[channel_name])
                except (ValueError, KeyError):
                    # Corrupted or malformed row — skip just this row
                    continue

                if math.isnan(value) or math.isinf(value):
                    continue

                values.append(value)

    except (OSError, csv.Error) as exc:
        print(f"Warning: could not read '{file_path}' ({exc}) — skipping file.")

    return values


def _calculate_month_counts(month_folder, thresholds, channel_name):
    """
    Computes the exceedance count for a single month, for every threshold,
    for the given channel_name.

    Returns a dict: {threshold: exceedance_count}
    """
    month_counts = {t: 0 for t in thresholds}

    attenuation_files = find_attenuation_files(month_folder)

    for file_path in attenuation_files:
        values = _read_channel_values(file_path, channel_name)

        for threshold in thresholds:
            exceed_count = sum(1 for v in values if v > threshold)
            month_counts[threshold] += exceed_count

    return month_counts


def calculate_monthly_exceedance(year_folder: str, channel_name: str, progress_callback=None):
    """
    Scans every month folder inside the year folder and computes, for
    every threshold, the number of 1 Hz samples where
        <channel_name> > Threshold
    separately for each month.

    Parameters
    ----------
    channel_name : str
        The attenuation column to analyze (e.g. "Att_Channel-1").
    progress_callback : callable, optional
        If provided, called as progress_callback(month_label, month_index,
        total_months) immediately after each month finishes processing,
        so a caller can update a live dashboard.

    Returns
    -------
    thresholds : list of float
        The sorted list of thresholds (LOWER_LIMIT..channel's upper limit).
    month_labels : list of str
        The month names found, in calendar order.
    counts : dict
        Nested dict: counts[month_label][threshold] = exceedance count (int)
    """
    thresholds = _build_thresholds(
        CHANNEL_UPPER_LIMITS.get(channel_name, DEFAULT_UPPER_LIMIT)
    )

    month_folders = find_months(year_folder)

    # Preserve calendar order, but only include months that were found
    found_labels = []
    counts = {}

    total_months = len(month_folders)

    for idx, month_folder in enumerate(month_folders, start=1):
        label = _month_label(month_folder.name)

        if label not in found_labels:
            found_labels.append(label)
            counts[label] = {t: 0 for t in thresholds}

        month_counts = _calculate_month_counts(month_folder, thresholds, channel_name)

        for threshold in thresholds:
            counts[label][threshold] += month_counts[threshold]

        if progress_callback is not None:
            progress_callback(label, idx, total_months)

    # Sort discovered month labels into calendar order
    month_labels = [m for m in MONTH_ORDER if m in found_labels]
    # Append any unrecognised labels (fallback case) at the end, as found
    month_labels += [m for m in found_labels if m not in MONTH_ORDER]

    return thresholds, month_labels, counts


# ==============================================================================
# TABLE BUILDING & DISPLAY
# ==============================================================================

def build_table(thresholds, month_labels, counts, upper_limit):
    """
    Builds the output table as a list of rows and a list of headers,
    ready to be passed to tabulate().

    Each row:
        [Lower Limit, Upper Limit, <month1 count>, ..., <monthN count>, Total Seconds]

    upper_limit is the channel-specific ceiling used to fill the
    "Upper Limit" column (each channel may have its own value).
    """
    headers = ["Lower Limit", "Upper Limit"] + month_labels + ["Total Seconds"]

    rows = []

    for threshold in thresholds:
        month_counts = [counts[label][threshold] for label in month_labels]
        total_seconds = sum(month_counts)

        row = [f"{threshold:.2f}", f"{upper_limit:.2f}"] + month_counts + [total_seconds]
        rows.append(row)

    return headers, rows


def print_table(headers, rows):
    """
    Prints the exceedance table using tabulate with a grid format.
    """
    print(tabulate(rows, headers=headers, tablefmt="grid"))


def _thin_border() -> Border:
    """Returns a Border object with thin lines on all four sides."""
    side = Side(style="thin")
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_header_style(cell) -> None:
    """Applies bold, centred, light-blue header formatting with a thin border."""
    cell.font      = Font(name="Arial", bold=True, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill      = PatternFill("solid", start_color="BDD7EE")
    cell.border    = _thin_border()


def _apply_data_style(cell) -> None:
    """Applies centred Arial formatting with a thin border to a data cell."""
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()


def _write_worksheet(ws, headers, rows) -> None:
    """
    Writes one exceedance table (headers + rows) into the given worksheet,
    using exactly the same formatting as before:
        • Bold, centre-aligned, light-blue header row
        • Thin borders on every cell
        • Numeric cells stored as numbers (not strings)
        • Auto-adjusted column widths
        • First row frozen

    This is the same logic that previously lived inline inside save_report —
    extracted unchanged so it can be reused once per channel worksheet.
    """
    # ── Build column headers ──────────────────────────────────────────────────
    # Replace plain "Lower Limit" / "Upper Limit" labels with labelled versions
    excel_headers = []
    for h in headers:
        if h == "Lower Limit":
            excel_headers.append("Lower Limit (dB)")
        elif h == "Upper Limit":
            excel_headers.append("Upper Limit (dB)")
        else:
            excel_headers.append(h)

    # Write header row
    for col_idx, label in enumerate(excel_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        _apply_header_style(cell)

    # Freeze the header row
    ws.freeze_panes = "A2"

    # ── Write data rows ───────────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            # Convert threshold strings ("1.00", "1.10" …) to float;
            # leave month counts and Total Seconds as int.
            if col_idx <= 2:                        # Lower / Upper Limit columns
                cell_value = float(value)
            elif col_idx == len(row_data):          # Total Seconds (last column)
                cell_value = int(value)
            else:                                   # Monthly counts
                cell_value = int(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            _apply_data_style(cell)

            # Format threshold columns to 2 decimal places
            if col_idx <= 2:
                cell.number_format = "0.00"

    # ── Auto-adjust column widths ─────────────────────────────────────────────
    for col_idx, header_label in enumerate(excel_headers, start=1):
        col_letter = get_column_letter(col_idx)

        # Measure the widest content in this column
        max_width = len(header_label)
        for row_idx in range(2, len(rows) + 2):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_width = max(max_width, len(str(cell_val)))

        ws.column_dimensions[col_letter].width = max_width + 4   # padding


def _channel_sheet_title(channel_name: str) -> str:
    """
    Converts an internal channel name like 'Att_Channel-1' into the
    worksheet title 'Channel-1' used in the workbook.
    """
    return channel_name.replace("Att_", "")


def save_report(
    year: str,
    channel_results: dict,
    output_root: str = "Processed_Data",
) -> Path:
    """
    Writes the exceedance report as a professionally formatted Excel workbook
    (.xlsx) inside <output_root>/Exceedance_Tables/Exceedance_Table_<year>.xlsx

    One worksheet is created per channel, named after the channel
    (e.g. "Channel-1", "Channel-3"), in the same order as TARGET_CHANNELS.
    Every worksheet uses the identical table format and formatting as before:

    Columns  : Lower Limit (dB) | Upper Limit (dB) | <months...> | Total Seconds
    Formatting:
        • Bold, centre-aligned, light-blue header row
        • Thin borders on every cell
        • Numeric cells stored as numbers (not strings)
        • Auto-adjusted column widths
        • First row frozen

    Parameters
    ----------
    channel_results : dict
        Ordered mapping of channel_name -> (headers, rows), one entry per
        channel in TARGET_CHANNELS order.

    Creates the Exceedance_Tables folder automatically if it does not exist.
    Returns the full Path of the saved workbook.
    """
    output_dir = Path(output_root) / "Exceedance_Tables"
    output_dir.mkdir(parents=True, exist_ok=True)

    report_path = output_dir / f"Exceedance_Table_{year}.xlsx"

    wb = Workbook()

    # openpyxl creates one default sheet — reuse it for the first channel,
    # then create additional sheets for every subsequent channel.
    default_ws = wb.active
    first = True

    for channel_name, (headers, rows) in channel_results.items():
        sheet_title = _channel_sheet_title(channel_name)

        if first:
            ws = default_ws
            ws.title = sheet_title
            first = False
        else:
            ws = wb.create_sheet(title=sheet_title)

        _write_worksheet(ws, headers, rows)

    wb.save(report_path)
    return report_path


# ==============================================================================
# TERMINAL DASHBOARD
# ==============================================================================

_CLEAR_LINE = "\033[2K"


def _clear_lines(n: int) -> None:
    """Moves the cursor up n lines and clears each one."""
    for _ in range(n):
        print("\033[A" + _CLEAR_LINE, end="")


def _bar(pct: float, width: int = 32) -> str:
    """Returns a filled/empty progress bar string for the given 0-100 pct."""
    filled = int(width * pct / 100)
    return "█" * filled + "-" * (width - filled)


class Dashboard:
    """
    Manages an in-place terminal dashboard for the exceedance engine,
    mirroring the style used in batch_processor.py.
    """

    def __init__(self, year: str):
        self.year = year
        self.scanned_months: list[str] = []
        self.progress_pct = 0
        self.stage = "scanning"   # "scanning" -> "generating" -> "saving" -> "done"
        self.saved_filename = ""
        self._lines_drawn = 0

        self._draw()

    # ------------------------------------------------------------------
    # Public interface
    # ------------------------------------------------------------------

    def month_scanned(self, month_label: str, idx: int, total: int) -> None:
        self.scanned_months.append(month_label)
        self.progress_pct = int(100 * idx / total) if total > 0 else 100
        self._draw()

    def set_stage(self, stage: str) -> None:
        self.stage = stage
        self._draw()

    def report_saved(self, filename: str) -> None:
        self.saved_filename = filename
        self.stage = "saved"
        self._draw()

    def close(self) -> None:
        self._draw()
        print()  # leave a trailing blank line after the dashboard settles

    # ------------------------------------------------------------------
    # Internal drawing helpers
    # ------------------------------------------------------------------

    def _draw(self) -> None:
        lines = self._build_lines()

        if self._lines_drawn > 0:
            _clear_lines(self._lines_drawn)

        for line in lines:
            print(line)

        self._lines_drawn = len(lines)

    def _build_lines(self) -> list:
        w = 58
        sep = "=" * w

        lines = []
        lines.append(sep)
        lines.append("DRSP Exceedance Statistics Engine")
        lines.append(sep)
        lines.append("")
        lines.append("Selected Year")
        lines.append(f"  {self.year}")
        lines.append("")
        lines.append("Scanning Months")
        for month in self.scanned_months:
            lines.append(f"  \u2713 {month}")
        if not self.scanned_months:
            lines.append("  (none yet)")
        lines.append("")

        if self.stage in ("generating", "saving", "saved"):
            lines.append("Generating Exceedance Table...")
            lines.append("Progress")
            lines.append(f"  [{_bar(self.progress_pct)}] {self.progress_pct}%")
            lines.append("")

        if self.stage in ("saving", "saved"):
            lines.append("Saving Report...")
            if self.stage == "saved":
                lines.append(f"  \u2713 {self.saved_filename}")
            lines.append("")

        lines.append(sep)
        if self.stage == "saved":
            lines.append("Completed Successfully")
            lines.append(sep)

        return lines


def _print_title() -> None:
    """Displays the engine's own title banner at startup."""
    width = 57
    print("=" * width)
    print("EXCEEDENCE ENGINE".center(width))
    print("=" * width)


# ==============================================================================
# MAIN
# ==============================================================================

def main():
    _print_title()

    start_time = time.monotonic()

    year_folder = select_year_folder()
    year = Path(year_folder).name

    dash = Dashboard(year)

    def _on_month_done(month_label, idx, total):
        dash.month_scanned(month_label, idx, total)

    # Automatically detect which attenuation channels exist in the selected
    # year's processed data — no dataset type is asked or hardcoded here.
    detected_channels = [
        channel
        for channel in detect_available_channels(year_folder)
        if channel in TARGET_CHANNELS
    ]

    if not detected_channels:
        dash.close()
        print("No attenuation channels detected. Nothing to report.")
        return

    # Process every detected channel through the SAME pipeline.
    # The dashboard is only wired to the FIRST channel's progress callback
    # so the user sees a single, unified job instead of repeated progress
    # output for each channel — subsequent channels reuse the identical
    # month-folder scan silently in the background.
    channel_results = {}
    total_rows = 0
    thresholds_count = 0
    any_months_found = False

    for i, channel_name in enumerate(detected_channels):
        callback = _on_month_done if i == 0 else None

        thresholds, month_labels, counts = calculate_monthly_exceedance(
            year_folder, channel_name, progress_callback=callback
        )

        if not month_labels:
            continue

        any_months_found = True
        upper_limit = CHANNEL_UPPER_LIMITS.get(channel_name, DEFAULT_UPPER_LIMIT)
        headers, rows = build_table(thresholds, month_labels, counts, upper_limit)
        channel_results[channel_name] = (headers, rows)
        total_rows += len(rows)
        thresholds_count = len(thresholds)

    if not any_months_found:
        dash.close()
        print("No month folders found. Nothing to report.")
        return

    dash.set_stage("generating")

    dash.set_stage("saving")
    report_path = save_report(year, channel_results)

    dash.report_saved(report_path.name)
    dash.close()

    elapsed = time.monotonic() - start_time

    print("\u2713 Report saved successfully")
    print()
    print("Location")
    print(f"  {report_path}")
    print()
    print("Channels Processed")
    print(f"  {', '.join(channel_results.keys())}")
    print()
    print("Rows Generated")
    print(f"  {total_rows}")
    print()
    print("Thresholds Processed")
    print(f"  {thresholds_count}")
    print()
    print("Execution Time")
    print(f"  {elapsed:.2f}s")


if __name__ == "__main__":
    main()
