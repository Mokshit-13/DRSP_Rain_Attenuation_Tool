"""
================================================================================
DRSP Rain Attenuation Tool — Data Validator
================================================================================
Standalone utility for inspecting the quality of raw NARL datasets.

Currently supports:
    • INF Value Validation

Architecture is designed for future extension.  Additional validation modules
(duplicate timestamps, missing timestamps, NaN values, invalid time formats,
etc.) can be added without redesigning the program — see ADDING VALIDATORS at
the bottom of this docstring.

This utility is read-only.  It NEVER modifies raw data, processed data, or any
other project file.
================================================================================
"""

from __future__ import annotations

import math
import sys
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog

import pandas as pd

from utils import find_main_narl_file


# ==============================================================================
# CONFIGURATION
# ==============================================================================

AMPLITUDE_CHANNELS = [
    "Amp_Channel-1",
    "Amp_Channel-2",
    "Amp_Channel-3",
    "Amp_Channel-4",
]

OUTPUT_ROOT = "Processed_Data"
REPORT_DIR  = "Validation_Reports"


# ==============================================================================
# TERMINAL HELPERS
# ==============================================================================

_SEP = "=" * 57


def _header(title: str) -> None:
    print()
    print(_SEP)
    print(title.center(57))
    print(_SEP)


def _clear_lines(n: int) -> None:
    """Moves the cursor up n lines and clears each one (in-place refresh)."""
    for _ in range(n):
        print("\033[A\033[2K", end="")


def _progress_bar(current: int, total: int, width: int = 28) -> str:
    filled = int(width * current / total) if total > 0 else 0
    bar    = "█" * filled + "-" * (width - filled)
    pct    = int(100 * current / total) if total > 0 else 0
    return f"[{bar}] {current}/{total} ({pct}%)"


def _render_progress(label: str, current: int, total: int, current_item: str) -> str:
    lines = [
        _SEP,
        "DATA VALIDATOR",
        _SEP,
        "",
        f"Scanning : {label}",
        "",
        f"  {_progress_bar(current, total)}",
        "",
        f"  \u23f3 {current_item}",
        "",
        _SEP,
    ]
    return "\n".join(lines)


# ==============================================================================
# UI — MENUS & FOLDER/FILE SELECTION
# ==============================================================================

def show_menu() -> str:
    """Displays the main menu and returns the user's validated choice."""
    _header("DATA VALIDATOR")
    print()
    print("  Select Validation Mode")
    print()
    print("    1.  Validate Single File")
    print("    2.  Validate Single Month")
    print("    3.  Validate Entire Year")
    print()
    print("    0.  Exit")
    print()

    while True:
        choice = input("  Enter Choice : ").strip()
        if choice in ("0", "1", "2", "3"):
            return choice
        print("  Invalid choice. Please enter 0, 1, 2, or 3.")


def select_report_format() -> str:
    """Displays the report format menu and returns '1' (TXT) or '2' (Excel)."""
    print()
    print(_SEP)
    print()
    print("  Select Report Format")
    print()
    print("    1.  TXT Report")
    print("    2.  Excel Report (.xlsx)")
    print()

    while True:
        choice = input("  Enter Choice : ").strip()
        if choice in ("1", "2"):
            return choice
        print("  Invalid choice. Please enter 1 or 2.")


def _open_folder_dialog(title: str) -> str:
    """Opens a standard Windows folder-selection dialog. Exits on cancel."""
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askdirectory(title=title)
    root.destroy()

    if not path:
        print("No folder selected. Exiting.")
        sys.exit(0)

    return path


def _open_file_dialog(title: str) -> str:
    """Opens a standard Windows file-selection dialog. Exits on cancel."""
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title=title,
        filetypes=[("Text files", "*.txt")],
    )
    root.destroy()

    if not path:
        print("No file selected. Exiting.")
        sys.exit(0)

    return path


# ==============================================================================
# INF DETECTION CORE
# ==============================================================================

def find_inf_timestamps(file_path) -> list[str]:
    """
    Reads a single raw NARL file and returns a sorted, deduplicated list of
    HH:MM:SS strings where at least one amplitude channel is ±inf.

    Rules
    -----
    • Each timestamp appears at most once, regardless of how many channels
      are infinite at that moment.
    • NaN values are ignored (only true infinity is flagged).
    • Unreadable files or missing columns return an empty list silently.
    """
    try:
        df = pd.read_csv(file_path, sep=r"\s+", engine="python")
    except Exception:
        return []

    present = [ch for ch in AMPLITUDE_CHANNELS if ch in df.columns]
    if not present or "Time" not in df.columns:
        return []

    for ch in present:
        df[ch] = pd.to_numeric(df[ch].astype(str).str.strip(), errors="coerce")

    inf_mask = df[present].apply(
        lambda col: col.apply(lambda v: isinstance(v, float) and math.isinf(v))
    ).any(axis=1)

    if not inf_mask.any():
        return []

    raw_times = df.loc[inf_mask, "Time"].astype(str).str.strip()
    cleaned   = raw_times.str.extract(r"(\d{1,2}:\d{2}:\d{2})", expand=False).dropna()

    return sorted(cleaned.unique().tolist())


# ==============================================================================
# SCANNING — returns structured result dicts
# ==============================================================================

def scan_file(file_path) -> dict:
    """
    Scans a single raw NARL file.

    Returns
    -------
    {
        "mode"       : "file",
        "source"     : str     — file path,
        "results"    : [{"folder": str, "timestamps": [str]}]
    }
    """
    file_path  = Path(file_path)
    timestamps = find_inf_timestamps(file_path)

    # Use the parent day-folder name as the report identifier, not the full path,
    # so the filename stays Windows-safe (no colons, backslashes, or slashes).
    day_folder = file_path.parent.name

    return {
        "mode"    : "file",
        "source"  : day_folder,
        "results" : [{"folder": file_path.name, "timestamps": timestamps}],
    }


def scan_month(month_folder, label: str = "", show_progress: bool = True) -> list[dict]:
    """
    Scans every daily subfolder inside a month folder (no rain-day filtering).

    Parameters
    ----------
    month_folder : str or Path
    label        : str  — label shown in the progress dashboard
    show_progress: bool — draw the in-place progress bar

    Returns
    -------
    list of {"folder": str, "timestamps": [str]}
    """
    month_path    = Path(month_folder)
    progress_label = label or month_path.name

    daily_folders = sorted(f for f in month_path.iterdir() if f.is_dir())
    total         = len(daily_folders)
    results       = []

    if show_progress:
        display = _render_progress(progress_label, 0, total, "")
        print(display)
        last_n = display.count("\n") + 1

    for i, folder in enumerate(daily_folders, start=1):
        if show_progress:
            _clear_lines(last_n)
            display = _render_progress(progress_label, i - 1, total, folder.name)
            print(display)
            last_n = display.count("\n") + 1

        timestamps = []
        try:
            main_file  = find_main_narl_file(folder)
            timestamps = find_inf_timestamps(main_file)
        except Exception:
            pass

        results.append({"folder": folder.name, "timestamps": timestamps})

    if show_progress:
        _clear_lines(last_n)
        display = _render_progress(progress_label, total, total, "Done")
        print(display)

    return results


def scan_year(year_folder) -> list[dict]:
    """
    Scans every month subfolder inside a year folder, scanning every daily
    folder within each month.

    Returns
    -------
    list of {"month": str, "folder": str, "timestamps": [str]}
    """
    year_path     = Path(year_folder)
    month_folders = sorted(f for f in year_path.iterdir() if f.is_dir())
    total_months  = len(month_folders)
    all_results   = []

    for idx, month_folder in enumerate(month_folders, start=1):
        month_name    = month_folder.name
        progress_label = f"{month_name}  [{idx}/{total_months}]"

        month_results = scan_month(
            month_folder,
            label=progress_label,
            show_progress=True,
        )

        for r in month_results:
            all_results.append({
                "month"      : month_name,
                "folder"     : r["folder"],
                "timestamps" : r["timestamps"],
            })

    return all_results


# ==============================================================================
# VALIDATION ENTRY POINTS
# ==============================================================================

def validate_single_file() -> dict:
    """Validates one raw NARL file selected by the user."""
    file_path = _open_file_dialog("Select Raw NARL Data File")
    return scan_file(file_path)


def validate_single_month() -> dict:
    """Validates every daily folder inside one month folder."""
    month_folder = _open_folder_dialog("Select Month Folder to Validate")
    month_name   = Path(month_folder).name
    results      = scan_month(month_folder, show_progress=True)

    return {
        "mode"       : "month",
        "source"     : month_name,
        "results"    : results,
    }


def validate_entire_year() -> dict:
    """Validates every daily folder inside every month of one year folder."""
    year_folder = _open_folder_dialog("Select Year Folder to Validate")
    year_name   = Path(year_folder).name
    results     = scan_year(year_folder)

    return {
        "mode"   : "year",
        "source" : year_name,
        "results": results,
    }


# ==============================================================================
# REPORT HELPERS
# ==============================================================================

def _report_path(stem: str, extension: str) -> Path:
    """Builds and creates the Validation_Reports output directory."""
    out_dir = Path(OUTPUT_ROOT) / REPORT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir / f"INF_Report_{stem}.{extension}"


def build_report_stem(data: dict) -> str:
    """
    Returns the filename stem (without extension) for any validation report.

    Rules
    -----
    Single File  → day-folder name with spaces replaced by underscores
                   e.g. "09122020 1Hz R"  →  "09122020_1Hz_R"
    Single Month → month folder name as-is
                   e.g. "December_2020"
    Entire Year  → year folder name as-is
                   e.g. "2020"
    """
    source = data["source"]

    if data["mode"] == "file":
        return source.replace(" ", "_")

    return source


def _summary_counts(data: dict) -> tuple[int, int, int]:
    """Returns (total_folders, affected_folders, total_occurrences)."""
    results    = data["results"]
    total      = len(results)
    affected   = sum(1 for r in results if r.get("timestamps"))
    occurrences = sum(len(r.get("timestamps", [])) for r in results)
    return total, affected, occurrences


def _mode_label(data: dict) -> str:
    mode = data["mode"]
    if mode == "file":
        return "Single File"
    if mode == "month":
        return "Single Month"
    if mode == "year":
        return "Entire Year"
    return mode


# ==============================================================================
# TXT REPORT
# ==============================================================================

def save_txt_report(data: dict) -> Path:
    """
    Generates a professionally formatted plain-text validation report.

    File is saved as:
        Processed_Data/Validation_Reports/INF_Report_<source>.txt
    """
    mode        = data["mode"]
    source      = data["source"]
    results     = data["results"]
    is_year     = (mode == "year")
    generated   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total, affected, occurrences = _summary_counts(data)

    report_path = _report_path(build_report_stem(data), "txt")

    width = 62

    with open(report_path, "w", newline="", encoding="utf-8") as f:

        def w(line: str = "") -> None:
            f.write(line + "\n")

        w("=" * width)
        w("INF VALUE VALIDATION REPORT".center(width))
        w("=" * width)
        w()
        w(f"Validation Mode  : {_mode_label(data)}")
        if mode == "file":
            w(f"File             : {source}")
        elif mode == "month":
            w(f"Month            : {source}")
        elif mode == "year":
            w(f"Year             : {source}")
        w(f"Generated On     : {generated}")
        w()
        w("=" * width)
        w()

        # Table
        affected_rows = [r for r in results if r.get("timestamps")]

        if not affected_rows:
            w("No INF values were found.")
        else:
            col_no     = 5
            col_month  = 18
            col_folder = 24
            col_times  = 24

            if is_year:
                div = (
                    f"+{'-' * (col_no + 2)}"
                    f"+{'-' * (col_month + 2)}"
                    f"+{'-' * (col_folder + 2)}"
                    f"+{'-' * (col_times + 2)}+"
                )
                hdr = (
                    f"| {'No.'.ljust(col_no)} "
                    f"| {'Month'.ljust(col_month)} "
                    f"| {'Day Folder'.ljust(col_folder)} "
                    f"| {'INF Occurrence Times'.ljust(col_times)} |"
                )
            else:
                div = (
                    f"+{'-' * (col_no + 2)}"
                    f"+{'-' * (col_folder + 2)}"
                    f"+{'-' * (col_times + 2)}+"
                )
                hdr = (
                    f"| {'No.'.ljust(col_no)} "
                    f"| {'Day Folder'.ljust(col_folder)} "
                    f"| {'INF Occurrence Times'.ljust(col_times)} |"
                )

            w(div)
            w(hdr)

            for idx, row in enumerate(affected_rows, start=1):
                folder  = row["folder"]
                times   = row["timestamps"]
                month   = row.get("month", "")

                w(div)

                for t_idx, t in enumerate(times):
                    no_str     = str(idx).ljust(col_no) if t_idx == 0 else "".ljust(col_no)
                    folder_str = folder[:col_folder].ljust(col_folder) if t_idx == 0 else "".ljust(col_folder)
                    time_str   = t.ljust(col_times)

                    if is_year:
                        month_str = month[:col_month].ljust(col_month) if t_idx == 0 else "".ljust(col_month)
                        w(f"| {no_str} | {month_str} | {folder_str} | {time_str} |")
                    else:
                        w(f"| {no_str} | {folder_str} | {time_str} |")

            w(div)

        w()
        w("=" * width)
        w("SUMMARY".center(width))
        w("=" * width)
        w()
        if is_year:
            months_scanned = len({r.get("month") for r in results})
            w(f"Months Scanned         : {months_scanned}")
        w(f"Daily Folders Scanned  : {total}")
        w(f"Folders Containing INF : {affected}")
        w(f"Total INF Occurrences  : {occurrences}")
        w()
        w("=" * width)

    return report_path


# ==============================================================================
# EXCEL REPORT
# ==============================================================================

def save_excel_report(data: dict) -> Path:
    """
    Generates a professionally formatted Excel workbook (.xlsx).

    Worksheet 1 — INF Validation : one row per timestamp occurrence
    Worksheet 2 — Summary        : metadata and counts

    File saved as:
        Processed_Data/Validation_Reports/INF_Report_<source>.xlsx
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    mode      = data["mode"]
    source    = data["source"]
    results   = data["results"]
    is_year   = (mode == "year")
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total, affected, occurrences = _summary_counts(data)

    report_path = _report_path(build_report_stem(data), "xlsx")

    # ── Style helpers ─────────────────────────────────────────────────────────
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hdr_fill = PatternFill("solid", start_color="BDD7EE")

    def _hdr_style(cell) -> None:
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill      = _hdr_fill
        cell.border    = _border

    def _data_style(cell, align: str = "center") -> None:
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = _border

    def _auto_width(ws, col_idx: int, values: list) -> None:
        col_letter = get_column_letter(col_idx)
        max_w      = max((len(str(v)) for v in values), default=8)
        ws.column_dimensions[col_letter].width = max_w + 4

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # ── Worksheet 1 : INF Validation ─────────────────────────────────────────
    ws1        = wb.active
    ws1.title  = "INF Validation"

    if is_year:
        headers = ["S.No", "Month", "Day Folder", "Time"]
    else:
        headers = ["S.No", "Day Folder", "Time"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=h)
        _hdr_style(cell)

    ws1.freeze_panes = "A2"

    row_num    = 2
    serial_no  = 1
    col_values = {h: [h] for h in headers}   # for auto-width measurement

    affected_rows = [r for r in results if r.get("timestamps")]

    if not affected_rows:
        # Write a single "no issues" row
        note_cell = ws1.cell(row=2, column=1, value="No INF values were found.")
        _data_style(note_cell, align="left")
        note_cell.font = Font(name="Arial", size=10, italic=True)
    else:
        for row in affected_rows:
            folder = row["folder"]
            month  = row.get("month", "")
            times  = row["timestamps"]

            for t in times:
                if is_year:
                    values = [serial_no, month, folder, t]
                else:
                    values = [serial_no, folder, t]

                for col_idx, val in enumerate(values, start=1):
                    cell = ws1.cell(row=row_num, column=col_idx, value=val)
                    _data_style(cell)
                    col_values[headers[col_idx - 1]].append(val)

                serial_no += 1
                row_num   += 1

    # Auto-adjust column widths
    for col_idx, h in enumerate(headers, start=1):
        _auto_width(ws1, col_idx, col_values.get(h, [h]))

    # ── Worksheet 2 : Summary ─────────────────────────────────────────────────
    ws2       = wb.create_sheet("Summary")
    s_headers = ["Field", "Value"]

    for col_idx, h in enumerate(s_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        _hdr_style(cell)

    ws2.freeze_panes = "A2"

    summary_rows = [
        ("Validation Mode",       _mode_label(data)),
    ]
    if mode == "file":
        summary_rows.append(("File", source))
    elif mode == "month":
        summary_rows.append(("Month", source))
    elif mode == "year":
        summary_rows.append(("Year", source))
        months_scanned = len({r.get("month") for r in results})
        summary_rows.append(("Months Scanned", months_scanned))

    summary_rows += [
        ("Generated On",           generated),
        ("Daily Folders Scanned",  total),
        ("Folders Containing INF", affected),
        ("Total INF Occurrences",  occurrences),
    ]

    for row_idx, (field, value) in enumerate(summary_rows, start=2):
        c_field = ws2.cell(row=row_idx, column=1, value=field)
        c_value = ws2.cell(row=row_idx, column=2, value=value)
        _data_style(c_field, align="left")
        _data_style(c_value, align="left")

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 36

    wb.save(report_path)
    return report_path


# ==============================================================================
# MAIN
# ==============================================================================

def main() -> None:
    choice = show_menu()

    if choice == "0":
        print()
        print("  Exiting.")
        print()
        sys.exit(0)

    fmt_choice = select_report_format()

    print()
    print(_SEP)
    print()

    # ── Run the selected validation ───────────────────────────────────────────
    if choice == "1":
        data = validate_single_file()
    elif choice == "2":
        data = validate_single_month()
    elif choice == "3":
        data = validate_entire_year()
    else:
        sys.exit(1)

    # ── Generate the selected report ──────────────────────────────────────────
    print()
    print("  Saving report...")

    if fmt_choice == "1":
        report_path = save_txt_report(data)
    else:
        report_path = save_excel_report(data)

    # ── Final terminal summary ────────────────────────────────────────────────
    total, affected, occurrences = _summary_counts(data)

    print()
    print(_SEP)
    print("Validation Completed Successfully".center(57))
    print(_SEP)
    print()
    print("  Report Saved")
    print(f"  {report_path}")
    print()
    print(f"  Daily Folders Scanned  : {total}")
    print(f"  Folders Containing INF : {affected}")
    print(f"  Total INF Occurrences  : {occurrences}")
    print()
    print(_SEP)
    print()


if __name__ == "__main__":
    main()